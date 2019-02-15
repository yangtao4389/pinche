"""
从file/params1.xlsx读取数据，手动写入到keywords.py里面的Series_KEY_MAP  Series_property_api
关于params1.xlsx 数据来源于接口文档

"""

# from openpyxl import load_workbook
# wb = load_workbook(filename='鉴权.xlsx', read_only=True)
# ws = wb['Sheet10']
#
# li = []
# key = []
# key_value = []
# item = {}
# for num_row,row in enumerate(ws.rows):
#     if num_row < 1 :
#         continue
#     for index,cell in enumerate(row):
#         # item = {}
#         if index == 0:
#             a = cell.value
#             key.append(a)
#             item[a] = None
#         if index == 1:
#             b = cell.value
#             li.append(cell.value)
#             item[a] = b
#
#             # key_value.append(item)
# print(key_value)
# print(key)
# print(item)


# 修改单元格里面的内容并保存
# import re
# from openpyxl import load_workbook
# from openpyxl.worksheet.write_only import WriteOnlyCell
# wb = load_workbook(filename='电竞媒资导入表--天津HD.xlsx')
# ws = wb['媒资信息HD']
#
# li = []
# key = []
# key_value = []
# item = {}
# for num_row,row in enumerate(ws.rows):
#     if num_row < 1 :
#         continue
#     for index,cell in enumerate(row):
#         # item = {}
#         if index == 3:
#             tmp = str(cell.value)
#             tmp_list = tmp.split("/")
#             tmp_value = tmp_list[-1]
#             # print(tmp_list)
#             # print(dir(ws))
#             key_value.append(item)
#             ws.cell(row=num_row+1,column=5).value = tmp_value
#
# # ws.cell(row=1, column=1).value = 'aaa'
# # cell = WriteOnlyCell(ws, value="hello world")
# # ws.append([cell])
# wb.save('write_only_file.xlsx')

# 写sql语句
def insertintoSQL():
    from resource_.models import ResourceCollection,ResourceCollectionVideo,ResourceVideo,ResourceKind,ResourceKindCollection
    from openpyxl import load_workbook
    from openpyxl.worksheet.write_only import WriteOnlyCell
    wb = load_workbook(filename='common\电竞媒资导入表 第二批 天津联通.xlsx')
    ws = wb['Sheet1']

    li = []
    key = []
    key_value = []
    item_video = {}
    item_collection = {}
    item_kind = {}
    for num_row,row in enumerate(ws.rows):
        if num_row < 1 :
            continue
        for index,cell in enumerate(row):
            if index == 0:
                print(cell.value, '1')
                item_video.update(dict(
                    CorrelateID = cell.value,
                    c_key = cell.value
                ))
            if index == 1:
                item_video.update(dict(
                    CDNPlayUrl=cell.value
                ))
            if index == 2:
                item_video.update(dict(
                    c_name=cell.value
                ))
            if index == 3:
                item_video.update(dict(
                    c_img="resource/%s"%cell.value
                ))
            if index == 4:
                item_collection.update(dict(
                    c_key = cell.value
                ))
            if index == 5:
                item_collection.update(dict(
                    c_name=cell.value
                ))
            if index == 6:
                item_collection.update(dict(
                    c_img_detail="resource/%s" % cell.value
                ))
            if index == 7:
                item_collection.update(dict(
                    c_img="resource/%s" % cell.value
                ))
            if index == 8:
                item_kind.update(dict(
                    c_key=cell.value
                ))
            if index == 9:
                item_kind.update(dict(
                    c_name=cell.value
                ))
                # item_video.update(dict(
                #     f_collection = ResourceCollection.objects.get(c_key=cell.value)
                # ))
        print(item_video,item_collection,item_kind)

        # 增加视频
        a,created = ResourceVideo.objects.get_or_create(CorrelateID=item_video["CorrelateID"])
        # if created:
        ResourceVideo.objects.filter(CorrelateID=item_video["CorrelateID"]).update(**item_video)

        # 增加专辑
        b,c2 = ResourceCollection.objects.get_or_create(c_key=item_collection["c_key"])
        # if c2:
        ResourceCollection.objects.filter(c_key=item_collection["c_key"]).update(**item_collection)

        # 增加分类
        c,c3 = ResourceKind.objects.get_or_create(c_key=item_kind["c_key"])
        # if c3:
        ResourceKind.objects.filter(c_key=item_kind["c_key"]).update(**item_kind)


        # 视频映射专辑
        ResourceCollectionVideo.objects.get_or_create(f_collection=b,f_video=a)

        # 分类映射专辑
        ResourceKindCollection.objects.get_or_create(f_kind=c,f_collection=b)

        # break

        # ResourceCollectionVideo.objects.get_or_create(**item)
        # break
        # ResourceCollection.objects.get_or_create(**item)
        # ResourceCollection(
        #     **item
        # ).save()
        # ResourceCollection.objects.filter(c_key=item["c_key"]).update(**item)


def insertinto_bq_SQL():
    from resourceSQ.models import ResourceSQCollection,ResourceSQCollectionVideo,ResourceSQVideo,ResourceSQKind,ResourceSQKindCollection
    from openpyxl import load_workbook
    from openpyxl.worksheet.write_only import WriteOnlyCell
    wb = load_workbook(filename='common\电竞媒资导入表--天津HD.xlsx')
    ws = wb['媒资信息SD']
    item_video = {}
    item_collection = {}
    item_kind = {}
    for num_row,row in enumerate(ws.rows):
        if num_row < 1 :
            continue
        for index,cell in enumerate(row):
            if index == 0:
                print(cell.value, '1')
                item_video.update(dict(
                    CorrelateID = cell.value,
                    c_key = cell.value
                ))
            if index == 1:
                item_video.update(dict(
                    CDNPlayUrl=cell.value
                ))
            if index == 2:
                item_video.update(dict(
                    c_name=cell.value
                ))
            if index == 3:
                item_video.update(dict(
                    c_img="resourceSQ/%s"%cell.value
                ))
            if index == 4:
                item_collection.update(dict(
                    c_key = cell.value
                ))
            if index == 5:
                item_collection.update(dict(
                    c_name=cell.value
                ))
            if index == 6:
                item_collection.update(dict(
                    c_img_detail="resourceSQ/%s" % cell.value
                ))
            if index == 7:
                item_collection.update(dict(
                    c_img="resourceSQ/%s" % cell.value
                ))
            if index == 8:
                item_kind.update(dict(
                    c_key=cell.value
                ))
            if index == 9:
                item_kind.update(dict(
                    c_name=cell.value
                ))

        print(item_video,item_collection,item_kind)

        # 增加视频
        # a,created = ResourceSQVideo.objects.get_or_create(CorrelateID=item_video["CorrelateID"])
        # if created:
        #     ResourceSQVideo.objects.filter(CorrelateID=item_video["CorrelateID"]).update(**item_video)

        # 增加专辑
        b,c2 = ResourceSQCollection.objects.get_or_create(c_key=item_collection["c_key"])
        # if c2:
        ResourceSQCollection.objects.filter(c_key=item_collection["c_key"]).update(**item_collection)

        # 增加分类
        # c,c3 = ResourceSQKind.objects.get_or_create(c_key=item_kind["c_key"])
        # if c3:
        #     ResourceSQKind.objects.filter(c_key=item_kind["c_key"]).update(**item_kind)
        #
        #
        # # 视频映射专辑
        # ResourceSQCollectionVideo.objects.get_or_create(f_collection=b,f_video=a)
        #
        # # 分类映射专辑
        # ResourceSQKindCollection.objects.get_or_create(f_kind=c,f_collection=b)


# insertintoSQL()
# insertinto_bq_SQL()
















